"""The actual reconciliation pipeline, shared by the CLI (cli.py) and the
Flask UI (webui.py) so neither has to duplicate this logic. Takes a plain
PipelineOptions (no argparse.Namespace, no Flask request object) and returns
a PipelineResult; all progress/status text goes into `logs` rather than being
printed directly, so callers decide how to surface it.
"""

import json
import os
from dataclasses import dataclass, field
from typing import Optional

from . import merge_utils, user_config
from .anomalies import detect_anomalies
from .matcher import match_all
from .pacing_sheet import (
    find_duplicate_campaign_names,
    group_by_goal_hits_merge,
    read_pacing_sheet_from_file,
    read_pacing_sheet_from_values,
)
from .parsers import parse_all_reports


@dataclass
class PipelineOptions:
    meta: Optional[str] = None
    linkedin: Optional[str] = None
    google_ads: Optional[str] = None
    taboola: Optional[str] = None
    stackadapt: Optional[str] = None

    pacing_sheet: Optional[str] = None
    pacing_tab: str = "Pacing sheet"

    spreadsheet_id: Optional[str] = None
    sheets_credentials: Optional[str] = None
    write_to_sheet: bool = False

    update_goal_hits: bool = False
    goal_hits_headless: bool = True
    goal_hits_limit: int = 0
    goal_hits_no_data_marker: Optional[str] = "N/A"
    dotenv_path: str = ".env"

    pressboard_api_key: Optional[str] = None

    output_dir: str = "./output"


@dataclass
class PipelineResult:
    pacing_row_count: int = 0
    duplicate_count: int = 0
    match_summary: dict = field(default_factory=dict)
    anomaly_flags: list = field(default_factory=list)
    read_errors: dict = field(default_factory=dict)
    sheet_write_result: Optional[dict] = None
    goal_hits_result: Optional[dict] = None
    logs: list = field(default_factory=list)


def load_pacing_rows(options: PipelineOptions) -> list:
    if options.pacing_sheet:
        return read_pacing_sheet_from_file(options.pacing_sheet, sheet_name=options.pacing_tab)
    if options.spreadsheet_id and options.sheets_credentials:
        from . import sheets_writer

        values = sheets_writer.read_pacing_values(options.spreadsheet_id, options.pacing_tab, options.sheets_credentials)
        return read_pacing_sheet_from_values(values)
    raise SystemExit("Provide either --pacing-sheet <file> or both --spreadsheet-id and --sheets-credentials to read the Pacing sheet.")


def serialize_match_result(result) -> dict:
    row = result.pacing_row
    entry = {
        "row_index": row.sheet_row - 1,
        "campaign_name": row.campaign_name,
        "platform": result.platform,
        "match_type": result.match_type,
        "matched_report_name": None,
        "spent": None,
        "impressions": None,
        "clicks": None,
        "reason": result.reason,
    }
    if result.matched is not None:
        entry["matched_report_name"] = result.matched.match_key
        entry["spent"] = result.matched.spent
        entry["impressions"] = result.matched.impressions
        entry["clicks"] = result.matched.clicks
        if result.matched.note:
            entry["note"] = result.matched.note
    if result.candidates:
        entry["candidates"] = [c.match_key for c in result.candidates]
    return entry


def build_match_summary(pacing_rows, match_results, duplicates, reports, read_errors) -> dict:
    by_platform = {}
    unmatched = []
    for result in match_results:
        platform = result.platform or "Unknown"
        stats = by_platform.setdefault(platform, {"matched": 0, "unmatched": 0, "ambiguous": 0, "no_report": 0})
        if result.match_type in ("exact", "contained", "id_match"):
            stats["matched"] += 1
        elif result.match_type == "ambiguous":
            stats["ambiguous"] += 1
            unmatched.append({"campaign_name": result.pacing_row.campaign_name, "reason": result.reason, "candidates": [c.match_key for c in result.candidates]})
        elif result.match_type == "no_report":
            stats["no_report"] += 1
            unmatched.append({"campaign_name": result.pacing_row.campaign_name, "reason": "no report uploaded for platform"})
        else:
            stats["unmatched"] += 1
            unmatched.append({"campaign_name": result.pacing_row.campaign_name, "reason": result.reason or "campaign not found in report"})

    matched_source_indices = {}
    for result in match_results:
        if result.platform is None:
            continue
        seen = matched_source_indices.setdefault(result.platform, set())
        if result.matched is not None:
            seen.add(result.matched.source_index)
        for cand in result.candidates:
            seen.add(cand.source_index)

    new_campaigns = []
    for platform, df in reports.items():
        seen = matched_source_indices.get(platform, set())
        unseen = df[~df["source_index"].isin(seen)]
        for _, r in unseen.iterrows():
            new_campaigns.append({"platform": platform, "campaign_name": r["match_key"]})

    return {
        "total_pacing_campaigns": len(pacing_rows),
        "by_platform": by_platform,
        "unmatched": unmatched,
        "new_campaigns_not_in_pacing_sheet": new_campaigns,
        "duplicate_campaign_names": duplicates,
        "report_read_errors": read_errors,
    }


def _write_json(path, payload) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, default=str)


def run_pipeline(options: PipelineOptions, on_progress=None) -> PipelineResult:
    logs = []
    result = PipelineResult(logs=logs)

    def notify(message):
        logs.append(message)
        if on_progress:
            on_progress(message)

    saved = user_config.load_config()
    if options.spreadsheet_id is None:
        options.spreadsheet_id = saved.get("spreadsheet_id")
    if options.pacing_tab == "Pacing sheet" and "pacing_tab" in saved:
        options.pacing_tab = saved["pacing_tab"]
    if options.sheets_credentials is None:
        options.sheets_credentials = saved.get("sheets_credentials")
    user_config.save_config(
        spreadsheet_id=options.spreadsheet_id,
        pacing_tab=options.pacing_tab,
        sheets_credentials=options.sheets_credentials,
    )

    os.makedirs(options.output_dir, exist_ok=True)

    notify("Reading Pacing sheet...")
    pacing_rows = load_pacing_rows(options)
    result.pacing_row_count = len(pacing_rows)
    duplicates = find_duplicate_campaign_names(pacing_rows)
    duplicate_names = set(duplicates.keys())
    result.duplicate_count = len(duplicate_names)
    clean_rows = [r for r in pacing_rows if r.campaign_name not in duplicate_names]

    notify("Parsing platform reports...")
    report_paths = {
        "Meta": options.meta,
        "LinkedIn": options.linkedin,
        "Google Ads": options.google_ads,
        "Taboola": options.taboola,
        "StackAdapt": options.stackadapt,
    }
    reports, read_errors = parse_all_reports(report_paths)
    result.read_errors = read_errors

    notify("Matching campaigns against the Pacing sheet...")
    match_results = match_all(clean_rows, reports)
    anomalies = detect_anomalies(match_results)

    updated_tracker_data = [serialize_match_result(r) for r in match_results]
    match_summary = build_match_summary(clean_rows, match_results, duplicates, reports, read_errors)
    result.match_summary = match_summary
    result.anomaly_flags = [a.__dict__ for a in anomalies]

    _write_json(os.path.join(options.output_dir, "updated_tracker_data.json"), updated_tracker_data)
    _write_json(os.path.join(options.output_dir, "match_summary.json"), match_summary)
    _write_json(os.path.join(options.output_dir, "anomaly_flags.json"), result.anomaly_flags)

    notify(f"Pacing rows: {len(pacing_rows)} (duplicates flagged and skipped: {len(duplicate_names)})")
    for platform, stats in match_summary["by_platform"].items():
        notify(f"  {platform}: {stats}")
    notify(f"Unmatched/needs-review: {len(match_summary['unmatched'])}")
    notify(f"New campaigns not in Pacing sheet: {len(match_summary['new_campaigns_not_in_pacing_sheet'])}")
    notify(f"Anomaly flags: {len(result.anomaly_flags)}")
    if read_errors:
        notify(f"Report read errors: {read_errors}")

    if options.pressboard_api_key:
        notify("Fetching Goal Hits via legacy Pressboard API key...")
        result.goal_hits_result = _run_legacy_pressboard(clean_rows, options.pressboard_api_key, options.output_dir, notify)

    if options.write_to_sheet:
        notify("Writing Spent/Impressions/Clicks to the live Google Sheet...")
        result.sheet_write_result = _write_to_live_sheet(options, match_results, notify)
    else:
        notify("(--write-to-sheet not set: no changes were written to any Google Sheet)")

    if options.update_goal_hits:
        notify("Starting Goal Hits update from Pressboard...")
        result.goal_hits_result = _run_goal_hits(options, pacing_rows, notify)

    notify("Done.")
    return result


def _run_legacy_pressboard(pacing_rows, api_key, output_dir, notify) -> dict:
    """The original Bearer-token-API path (pressboard.py) -- kept for
    backward compatibility, but superseded by --update-goal-hits (the real
    StudioStack integration, see goal_hits.py/studiostack.py)."""
    from . import pressboard

    results = []
    for row in pacing_rows:
        try:
            value = pressboard.fetch_goal_hits(row.campaign_name, api_key)
        except Exception as exc:  # noqa: BLE001
            results.append({"campaign_name": row.campaign_name, "error": str(exc)})
            continue
        if value is not None:
            results.append({"campaign_name": row.campaign_name, "goal_hits": value})
    _write_json(os.path.join(output_dir, "goal_hits_report.json"), results)
    updated = sum("goal_hits" in r for r in results)
    notify(f"Pressboard (legacy API key path): fetched Goal Hits for {updated} campaign(s) (reference only -- not written to the sheet).")
    return {"legacy_updated": updated}


def _write_to_live_sheet(options: PipelineOptions, match_results, notify) -> dict:
    if not (options.spreadsheet_id and options.sheets_credentials):
        notify("ERROR: --write-to-sheet requires both --spreadsheet-id and --sheets-credentials.")
        return {"error": "missing spreadsheet_id/sheets_credentials"}

    from . import sheets_writer

    updates = []
    for result in match_results:
        if result.matched is None:
            continue
        updates.append(
            {
                "sheet_row": result.pacing_row.sheet_row,
                "spent": result.matched.spent,
                "impressions": result.matched.impressions,
                "clicks": result.matched.clicks,
            }
        )

    response = sheets_writer.write_updates(options.spreadsheet_id, options.pacing_tab, updates, options.sheets_credentials)
    total_cells = response.get("totalUpdatedCells", "n/a")
    notify(f"Wrote {len(updates)} row(s) ({total_cells} cells) to the live Pacing sheet.")
    return {"rows_written": len(updates), "cells_written": total_cells}


def _run_goal_hits(options: PipelineOptions, pacing_rows, notify) -> dict:
    from . import goal_hits, sheets_writer

    live = bool(options.spreadsheet_id and options.sheets_credentials and not options.pacing_sheet)

    if live:
        from .config import PACING_COLUMNS_0INDEXED

        merges = sheets_writer.get_merges(options.spreadsheet_id, options.pacing_tab, options.sheets_credentials)
        merge_map = merge_utils.build_merge_map_from_sheets_api(merges, PACING_COLUMNS_0INDEXED["goal_hits"])
    else:
        import openpyxl
        from openpyxl.utils import get_column_letter

        from .config import PACING_COLUMNS_0INDEXED

        wb = openpyxl.load_workbook(options.pacing_sheet, data_only=True)
        ws = wb[options.pacing_tab]
        goal_hits_letter = get_column_letter(PACING_COLUMNS_0INDEXED["goal_hits"] + 1)
        merge_map = merge_utils.build_merge_map_from_openpyxl(ws, goal_hits_letter)

    groups, flags = group_by_goal_hits_merge(pacing_rows, merge_map)
    if flags:
        notify(f"Goal Hits: {len(flags)} data-quality flag(s) (blank merge top row / inconsistent Goal) -- see goal_hits_flags.json")
        _write_json(os.path.join(options.output_dir, "goal_hits_flags.json"), flags)

    updates, log_rows = goal_hits.run_goal_hits_update(
        groups,
        headless=options.goal_hits_headless,
        limit=options.goal_hits_limit,
        dotenv_path=options.dotenv_path,
        no_data_marker=options.goal_hits_no_data_marker,
        on_progress=notify,
    )
    goal_hits.write_log(log_rows, os.path.join(options.output_dir, "goal_hits_update_log.csv"))

    written = 0
    if live and updates:
        response = sheets_writer.write_updates(options.spreadsheet_id, options.pacing_tab, updates, options.sheets_credentials)
        written = response.get("totalUpdatedCells", len(updates))
        notify(f"Goal Hits: wrote {len(updates)} cell(s) to the live Pacing sheet.")
    elif updates:
        written = _write_goal_hits_to_local_file(options.pacing_sheet, options.pacing_tab, updates)
        notify(f"Goal Hits: wrote {written} cell(s) to local file {options.pacing_sheet}.")
    else:
        notify("Goal Hits: no values found to write (see goal_hits_update_log.csv for per-group status).")

    updated_count = sum(1 for row in log_rows if row.get("status") == "updated")
    notify(f"Goal Hits: processed {len(log_rows)} group(s), {updated_count} updated, {len(log_rows) - updated_count} left as-is.")

    return {
        "groups_processed": len(log_rows),
        "groups_updated": updated_count,
        "cells_written": written,
        "data_quality_flags": len(flags),
    }


def _write_goal_hits_to_local_file(pacing_sheet_path, pacing_tab, updates) -> int:
    import openpyxl
    from openpyxl.utils import get_column_letter

    from .config import PACING_COLUMNS_0INDEXED

    goal_hits_letter = get_column_letter(PACING_COLUMNS_0INDEXED["goal_hits"] + 1)
    wb = openpyxl.load_workbook(pacing_sheet_path)
    ws = wb[pacing_tab]
    for update in updates:
        cell = f"{goal_hits_letter}{update['sheet_row']}"
        value = update["goal_hits"]
        ws[cell] = value if value != "" else None
    wb.save(pacing_sheet_path)
    return len(updates)
